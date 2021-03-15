from collections import defaultdict, Counter
import pandas
from openpyxl import load_workbook

# читаем Логи и создаём словарь
logs_data = pandas.read_excel('logs.xlsx', sheet_name='log', engine='openpyxl')
logs_data_dict = logs_data.to_dict('records')

# Создаём список товаров, разделяем его по запятым и удаляем лишние элементы
sales_list = []

# Определяем структуру записи словаря по продажам
#item_list = []
for elements in logs_data_dict:
    item_list = []
    sales_dict = {
    'gender': '',
    'month': 0,
    'items': [],
    }
    sales_dict['gender'] = elements['Пол']
    sales_dict['month'] = elements['Дата посещения'].month
    item_list.extend(elements['Купленные товары'].split(','))
    for del_sales in item_list:
        if del_sales == 'Ещё 2 варианта' or del_sales == 'Ещё 3 варианта':
            item_list.remove(del_sales)
    sales_dict['items'] = item_list
    sales_list.append(sales_dict)
#print(sales_list)

sales_dict = defaultdict(int)
count_sales = 0
while count_sales < len(sales_list):
    for element in sales_list[count_sales]['items']:
        sales_dict[element] += 1
    count_sales += 1
sales_counter = Counter(sales_dict)
#print(sales_counter)

# определяем количество месяцев в логе
max_month = 1
for element in logs_data_dict:
    if element['Дата посещения'].month > max_month:
        max_month = element['Дата посещения'].month

# Разбиваем статистику товаров по месяцам
sales_list_count = []
month = 1
while month <= max_month:
    sales_dict1 = defaultdict(int)
    for element in sales_list:
        if element['month'] == month:
            count_sales = 0
            while count_sales < len(element['items']):
                i = 0
                while i < 7:
                    if  element['items'][count_sales] == sales_counter.most_common(7)[i][0]:
                        sales_dict1[element['items'][count_sales]] += 1
                    i += 1
                count_sales += 1
    sales_counter1 = Counter(sales_dict1)
    sales_list_count.append(sales_counter1)
    month +=1
#print(sales_list_count[1].most_common(7)[0][0])
#print(sales_list_count[1].most_common(7)[0][1])

# Ищем самый популярный браузер
browser_dict = defaultdict(int)
for element in logs_data_dict:
    browser_dict[element['Браузер']] += 1
browser_counter = Counter(browser_dict)

# Разбиваем статистику браузеров по месяцам
browser_list = []
month = 1
while month <= max_month:
    browser_dict1 = defaultdict(int)
    for element in logs_data_dict:
        if element['Дата посещения'].month == month:
            i = 0
            while i < 7:
                if element['Браузер'] == browser_counter.most_common(7)[i][0]:
                    browser_dict1[element['Браузер']] += 1
                i += 1
    browser_counter1 = Counter(browser_dict1)
    browser_list.append(browser_counter1)
    month +=1

# статистика по полу покупателей
gend_m_list = []
gend_f_list = []
for element in sales_list:
    if element['gender'] == 'м':
        gend_m_list.append(element)
    else:
        gend_f_list.append(element)

sales_m_dict = defaultdict(int)
count_sales = 0
while count_sales < len(gend_m_list):
    for element in gend_m_list[count_sales]['items']:
        sales_m_dict[element] += 1
    count_sales += 1
sales_m_counter = Counter(sales_m_dict)

sales_f_dict = defaultdict(int)
count_sales = 0
while count_sales < len(gend_f_list):
    for element in gend_f_list[count_sales]['items']:
        sales_f_dict[element] += 1
    count_sales += 1
sales_f_counter = Counter(sales_f_dict)

# выводим в файл 7 наиболее популярных браузеров по месяцам в отчёт
wb = load_workbook(filename='report.xlsx')
sheet = wb['Лист1'];

i = 0
#print(max_month)
while i < 7:
    m = 1
    sheet.cell(row=i+5, column=1).value = f"{browser_counter.most_common(7)[i][0]}"
    sheet.cell(row=i+19, column=1).value = f"{sales_counter.most_common(7)[i][0]}"
    #print(sales_counter.most_common(7)[i][0])
    #print(sales_counter.most_common(7)[i][1])
    while m <= max_month:
        #print(f'{browser_counter.most_common(7)[i][0]} - {browser_list[m].most_common(7)[i][1]}')
        sheet.cell(row=i+5, column=m+2).value = f"{browser_list[m-1].most_common(7)[i][1]}"
        sheet.cell(row=i+19, column=m+2).value = f"{sales_list_count[m-1].most_common(7)[i][1]}"
        m += 1
    i += 1
sheet.cell(row=31, column=2).value = f"{sales_m_counter.most_common(1)[0][0]}"
sheet.cell(row=32, column=2).value = f"{sales_f_counter.most_common(1)[0][0]}"
sheet.cell(row=33, column=2).value = f"{sales_m_counter.most_common()[len(sales_m_counter)-1][0]}"
sheet.cell(row=34, column=2).value = f"{sales_f_counter.most_common()[len(sales_f_counter)-1][0]}"
wb.save('report.xlsx')
# print(logs_data_dict)